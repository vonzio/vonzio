#!/usr/bin/env python3
"""Render a spreadsheet to a self-contained HTML document for the dashboard's
Document deck tab (#368). Values only (openpyxl read_only) — no formulas, no
charts. Every sheet becomes a section; oversized sheets are truncated with a
visible notice rather than silently cut.

The output is displayed inside a sandboxed iframe (no scripts), so this emits
plain markup + inline CSS only. All cell content is HTML-escaped.

Usage: xlsx2html.py <src.xlsx|xls|ods> <out.html>
"""
import html
import sys

MAX_ROWS = 2000
MAX_COLS = 100

STYLE = """
body { font-family: -apple-system, 'Segoe UI', Roboto, sans-serif; margin: 16px;
       color: #1a1a1a; background: #fff; }
h2 { font-size: 14px; margin: 20px 0 8px; }
table { border-collapse: collapse; font-size: 12.5px; margin-bottom: 8px; }
td, th { border: 1px solid #d8d8d8; padding: 3px 8px; text-align: left;
         max-width: 360px; overflow: hidden; text-overflow: ellipsis;
         white-space: nowrap; }
th { background: #f3f3f3; font-weight: 600; }
.truncated { color: #8a6d1a; font-size: 12px; margin: 4px 0 12px; }
.empty { color: #888; font-size: 12.5px; }
@media (prefers-color-scheme: dark) {
  body { color: #e8e8e8; background: #1b1b1f; }
  td, th { border-color: #3a3a40; }
  th { background: #26262c; }
}
"""


def cell_text(value) -> str:
    if value is None:
        return ""
    return html.escape(str(value))


def main() -> int:
    if len(sys.argv) != 3:
        print("usage: xlsx2html.py <src> <out.html>", file=sys.stderr)
        return 2
    src, out_path = sys.argv[1], sys.argv[2]

    if src.lower().endswith(".ods"):
        # openpyxl doesn't read ODS; LibreOffice converts it to xlsx first.
        import subprocess, tempfile, os, glob
        tmp = tempfile.mkdtemp(prefix="ods2xlsx-")
        subprocess.run(
            ["soffice", f"-env:UserInstallation=file://{tmp}/lo", "--headless",
             "--norestore", "--convert-to", "xlsx", "--outdir", tmp, src],
            check=True, capture_output=True,
        )
        converted = glob.glob(os.path.join(tmp, "*.xlsx"))
        if not converted:
            print("ods→xlsx conversion produced no output", file=sys.stderr)
            return 4
        src = converted[0]

    from openpyxl import load_workbook

    wb = load_workbook(src, read_only=True, data_only=True)
    parts = ["<!doctype html><meta charset='utf-8'>",
             f"<style>{STYLE}</style>"]
    for ws in wb.worksheets:
        parts.append(f"<h2>{html.escape(ws.title)}</h2>")
        rows_out = []
        truncated_rows = truncated_cols = False
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r >= MAX_ROWS:
                truncated_rows = True
                break
            if len(row) > MAX_COLS:
                truncated_cols = True
                row = row[:MAX_COLS]
            tag = "th" if r == 0 else "td"
            cells = "".join(f"<{tag}>{cell_text(v)}</{tag}>" for v in row)
            rows_out.append(f"<tr>{cells}</tr>")
        if rows_out:
            parts.append(f"<table>{''.join(rows_out)}</table>")
        else:
            parts.append("<p class='empty'>(empty sheet)</p>")
        if truncated_rows or truncated_cols:
            what = []
            if truncated_rows:
                what.append(f"first {MAX_ROWS} rows")
            if truncated_cols:
                what.append(f"first {MAX_COLS} columns")
            parts.append(
                f"<p class='truncated'>Preview truncated to the {' and '.join(what)} — "
                "download the file for the full data.</p>")
    wb.close()

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("".join(parts))
    return 0


if __name__ == "__main__":
    sys.exit(main())
